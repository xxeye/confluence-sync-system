"""
NoteLoader
讀取 xlsx 說明文件，提供圖片說明對照表給 page_builder 使用。

xlsx 格式：
  A 欄：檔名 或 群組名（不含副檔名亦可）
  B 欄：說明文字

對應規則：
  - 一般圖片  → key = 完整檔名，如 main_bg.png
  - 多國群組  → key = 群組名，如 main_btn_start（不含語系與副檔名）
  - NU 群組   → key = 群組名，如 main_nu_win（不含序號與副檔名）
"""

from pathlib import Path
from typing import Dict, Optional


class NoteLoader:
    """讀取 xlsx 說明文件，回傳 {key: note} 對照表"""

    def __init__(self, notes_file: Optional[str]):
        """
        Args:
            notes_file: xlsx 路徑（來自 config）。
                        為 None 或檔案不存在時靜默略過，回傳空 dict。
        """
        self._notes: Dict[str, str] = {}
        self._file = Path(notes_file) if notes_file else None
        self._load()

    def reload(self) -> bool:
        """重新從磁碟讀取 xlsx（監聽模式下 xlsx 更新後呼叫）。
        
        Returns:
            True = 有成功載入新資料；False = 檔案不存在或讀取失敗（舊資料保留不變）
        """
        old_count = len(self._notes)
        new_notes: Dict[str, str] = {}

        if not self._file or not self._file.exists():
            return False

        try:
            import openpyxl
            wb = openpyxl.load_workbook(self._file, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=1, values_only=True):
                if not row or row[0] is None:
                    continue
                key = str(row[0]).strip()
                note = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                if key:
                    new_notes[key] = note
            wb.close()

            self._notes = new_notes
            print(f"[NoteLoader] 重新載入說明文件：{self._file}，共 {len(self._notes)} 筆（原 {old_count} 筆）")
            return True

        except ImportError:
            print("[NoteLoader] 缺少 openpyxl，無法重新載入。")
            return False
        except Exception as e:
            print(f"[NoteLoader] 重新載入失敗：{e}，保留舊資料。")
            return False

    def _load(self) -> None:
        """讀取 xlsx；失敗時靜默略過（不中斷同步流程）"""
        if not self._file:
            return
        if not self._file.exists():
            print(f"[NoteLoader] 說明文件不存在，略過：{self._file}")
            return

        try:
            import openpyxl
            wb = openpyxl.load_workbook(self._file, read_only=True, data_only=True)
            ws = wb.active

            for row in ws.iter_rows(min_row=1, values_only=True):
                if not row or row[0] is None:
                    continue
                key = str(row[0]).strip()
                note = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                if key:
                    self._notes[key] = note

            wb.close()
            print(f"[NoteLoader] 載入說明文件：{self._file}，共 {len(self._notes)} 筆")

        except ImportError:
            print("[NoteLoader] 缺少 openpyxl，請執行 pip install openpyxl。略過說明文件。")
        except Exception as e:
            print(f"[NoteLoader] 讀取說明文件失敗：{e}。略過說明文件。")

    def get(self, key: str) -> str:
        """
        取得說明文字。

        查詢順序：
          1. 完整 key（如 main_bg.png 或群組名 main_btn_start）
          2. 去掉副檔名後的 key（如 main_bg.png → main_bg）
          3. 找不到回傳空字串

        Args:
            key: 檔名或群組名

        Returns:
            說明文字，找不到則回傳 ""
        """
        if key in self._notes:
            return self._notes[key]

        # 嘗試去掉副檔名
        stem = Path(key).stem
        return self._notes.get(stem, "")

    def as_dict(self) -> Dict[str, str]:
        """回傳所有說明資料的副本（公開介面，避免外部直接存取 _notes）"""
        return dict(self._notes)

    def is_empty(self) -> bool:
        """是否沒有任何說明資料"""
        return len(self._notes) == 0

